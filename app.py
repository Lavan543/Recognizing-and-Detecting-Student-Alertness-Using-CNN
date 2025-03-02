from flask import Flask, render_template, request, jsonify
import os
import cv2
import face_recognition
import pickle
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import dlib
from tensorflow.keras.models import Sequential, load_model
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.keras.optimizers import Adam
import numpy as np

app = Flask(__name__)

# Directory to save student data
students_dir = 'AMajorProject'
student_images_dir = os.path.join(students_dir, 'student_images')
attendance_dir = os.path.join(students_dir, 'CNN_dataset')
active_dir = os.path.join(attendance_dir, 'active')
not_active_dir = os.path.join(attendance_dir, 'not_active')

# Ensure directories exist
os.makedirs(student_images_dir, exist_ok=True)
os.makedirs(attendance_dir, exist_ok=True)
os.makedirs(active_dir, exist_ok=True)
os.makedirs(not_active_dir, exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            student_name = request.form['name'].replace(" ", "_")
            camera = cv2.VideoCapture(0)

            while True:
                ret, frame = camera.read()
                if not ret:
                    break

                cv2.imshow("Press 's' to save image, 'q' to quit", frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord('s'):
                    image_path = os.path.join(student_images_dir, f"{student_name}.jpg")
                    cv2.imwrite(image_path, frame)
                    student_data = {'name': student_name, 'image_path': image_path}
                    with open(os.path.join(students_dir, f"{student_name}.pkl"), 'wb') as f:
                        pickle.dump(student_data, f)
                    camera.release()
                    cv2.destroyAllWindows()
                    return f"Student {student_name} registered successfully."
                elif key == ord('q'):
                    break

            camera.release()
            cv2.destroyAllWindows()
            return "Registration process exited."

        except Exception as e:
            return f"An error occurred: {e}"
    return render_template('register.html')
@app.route('/recognize', methods=['GET'])
def recognize():
    try:
        def load_student_encodings():
            student_files = [f for f in os.listdir(student_images_dir) if f.endswith('.jpg')]
            students = {}
            for student_file in student_files:
                student_name = os.path.splitext(student_file)[0]
                image = face_recognition.load_image_file(os.path.join(student_images_dir, student_file))
                encodings = face_recognition.face_encodings(image)
                if encodings:
                    encoding = encodings[0]
                    students[student_name] = {'encoding': encoding}
            return students

        def record_attendance(student_name):
            file_path = 'attendance.xlsx'
            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['Student Name', 'In Time', 'Out Time'])

            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == student_name:
                    last_time_str = row[1]
                    last_time = datetime.strptime(last_time_str, "%Y-%m-%d %H:%M:%S")
                    if now - last_time < timedelta(hours=6):
                        return f"{student_name} already recorded within the last 6 hours."

            sheet.append([student_name, current_time, ''])
            workbook.save(file_path)
            return f"Attendance recorded for {student_name} at {current_time}"

        detector = dlib.get_frontal_face_detector()
        predictor = dlib.shape_predictor("shape_predictor_68_face_landmarks.dat")
        face_rec_model = dlib.face_recognition_model_v1("dlib_face_recognition_resnet_model_v1.dat")

        students = load_student_encodings()
        if not students:
            return "No registered students found."

        camera = cv2.VideoCapture(0)
        recognized_name = "Unknown"
        while True:
            ret, frame = camera.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector(gray)

            for face in faces:
                shape = predictor(gray, face)
                face_descriptor = np.array(face_rec_model.compute_face_descriptor(frame, shape))
                matches = face_recognition.compare_faces([student['encoding'] for student in students.values()], face_descriptor)

                if True in matches:
                    first_match_index = matches.index(True)
                    recognized_name = list(students.keys())[first_match_index]
                    record_attendance(recognized_name)
                else:
                    recognized_name = "Unregistered"

                # Draw a rectangle around the face and put the recognized name on the frame
                (x, y, w, h) = (face.left(), face.top(), face.width(), face.height())
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, recognized_name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)

            cv2.imshow("Press 'q' to quit", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        camera.release()
        cv2.destroyAllWindows()
        return f"Recognized Student: {recognized_name}"

    except Exception as e:
        return f"An error occurred: {e}"


@app.route('/detect-alertness', methods=['GET'])
def detect_alertness():
    try:
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            return "Error: Unable to open camera."

        active_count = 0
        not_active_count = 0
        active_samples = []
        not_active_samples = []

        while active_count < 20 or not_active_count < 20:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            if len(faces) > 0:
                status = "Active"
                if active_count < 20:
                    active_samples.append(frame)
                    active_count += 1
            else:
                status = "Not Active"
                if not_active_count < 20:
                    not_active_samples.append(frame)
                    not_active_count += 1

            cv2.putText(frame, status, (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.imshow('Camera', frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        # Save collected samples
        for i, img in enumerate(active_samples):
            cv2.imwrite(os.path.join(active_dir, f"active_{i}.jpg"), img)
        for i, img in enumerate(not_active_samples):
            cv2.imwrite(os.path.join(not_active_dir, f"not_active_{i}.jpg"), img)

        cap.release()
        cv2.destroyAllWindows()
        return "Alertness detection complete."

    except Exception as e:
        return f"An error occurred: {e}"


from flask import Flask, render_template, request, jsonify, send_file
import os
import cv2
import numpy as np
import matplotlib.pyplot as plt
from tensorflow.keras.models import Sequential, load_model
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense
from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.keras.optimizers import Adam
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
import seaborn as sns
model_path = os.path.join(students_dir, 'student_alertness_model.h5')

def create_model():
    model = Sequential([
        Conv2D(32, (3, 3), activation='relu', input_shape=(64, 64, 3)),
        MaxPooling2D(2, 2),
        Conv2D(64, (3, 3), activation='relu'),
        MaxPooling2D(2, 2),
        Flatten(),
        Dense(128, activation='relu'),
        Dense(1, activation='sigmoid')  # Binary classification
    ])
    model.compile(optimizer=Adam(learning_rate=0.001), loss='binary_crossentropy', metrics=['accuracy'])
    return model

@app.route('/train', methods=['GET'])
def train_model():
    datagen = ImageDataGenerator(rescale=1./255, validation_split=0.2)
    train_generator = datagen.flow_from_directory(
        attendance_dir, target_size=(64, 64), batch_size=32, class_mode='binary', subset='training')
    val_generator = datagen.flow_from_directory(
        attendance_dir, target_size=(64, 64), batch_size=32, class_mode='binary', subset='validation')
    
    model = create_model()
    history = model.fit(train_generator, validation_data=val_generator, epochs=10)
    model.save(model_path)
    
    # Plot accuracy and loss
    plt.figure(figsize=(10, 5))
    plt.subplot(1, 2, 1)
    plt.plot(history.history['accuracy'], label='Train Accuracy')
    plt.plot(history.history['val_accuracy'], label='Val Accuracy')
    plt.legend()
    plt.title('Training & Validation Accuracy')
    
    plt.subplot(1, 2, 2)
    plt.plot(history.history['loss'], label='Train Loss')
    plt.plot(history.history['val_loss'], label='Val Loss')
    plt.legend()
    plt.title('Training & Validation Loss')
    plt.savefig('static/training_results.png')
    
    return jsonify({"message": "Model trained successfully!", "image": "static/training_results.png"})

@app.route('/evaluate', methods=['GET'])
def evaluate_model():
    if not os.path.exists(model_path):
        return jsonify({"error": "Model not found. Train the model first!"})
    
    model = load_model(model_path)
    datagen = ImageDataGenerator(rescale=1./255)
    test_generator = datagen.flow_from_directory(attendance_dir, target_size=(64, 64), batch_size=32, class_mode='binary', shuffle=False)
    
    y_true = test_generator.classes
    y_pred = (model.predict(test_generator) > 0.5).astype(int)
    
    accuracy = accuracy_score(y_true, y_pred)
    precision = precision_score(y_true, y_pred)
    recall = recall_score(y_true, y_pred)
    f1 = f1_score(y_true, y_pred)
    conf_matrix = confusion_matrix(y_true, y_pred)
    
    # Plot confusion matrix
    plt.figure(figsize=(5, 5))
    sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues', xticklabels=['Not Active', 'Active'], yticklabels=['Not Active', 'Active'])
    plt.xlabel('Predicted')
    plt.ylabel('Actual')
    plt.title('Confusion Matrix')
    plt.savefig('static/confusion_matrix.png')
    
    return jsonify({
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1,
        "image": "static/confusion_matrix.png"
    })

if __name__ == '__main__':
    app.run(debug=True)
